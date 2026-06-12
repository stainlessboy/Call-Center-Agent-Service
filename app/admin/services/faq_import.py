"""Parse FAQ xlsx files and import into the FaqItem table."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from sqlalchemy import delete

from app.db.models import FaqItem
from app.db.session import get_session

HEADER_ALIASES = {
    "question": {
        "question",
        "questions",
        "q",
        "faq question",
        "вопрос",
        "вопросы",
        "вопрос/проблема",
        "вопросы/проблемы",
        "вопросы по кредитам",
        "вопросы по кредитам?",
    },
    "answer": {
        "answer",
        "answers",
        "a",
        "response",
        "ответ",
        "ответы",
        "ответ/решение",
        "ответы/решения",
    },
}

SHEET_LANGUAGE_ALIASES = {
    "ru": {"ru", "russian", "русский", "рус", "российский"},
    "en": {"en", "eng", "english", "английский", "англ"},
    "uz": {"uz", "uzb", "uzbek", "uzbekskiy", "узбекский", "узб"},
}

# Wide multilingual format: one row = one FAQ item, per-language columns.
# Produced by the /admin/seed export (xlsx/csv) and by the SQLAdmin list
# export (column names or "Вопрос (RU)"-style labels).
WIDE_FIELDS = (
    "question_ru", "answer_ru",
    "question_en", "answer_en",
    "question_uz", "answer_uz",
)

_WIDE_HEADER_ALIASES: dict[str, str] = {}
for _field in WIDE_FIELDS:
    _kind, _lang = _field.split("_")
    _ru_word = "вопрос" if _kind == "question" else "ответ"
    for _alias in (_field, f"{_kind} {_lang}", f"{_ru_word} {_lang}"):
        _WIDE_HEADER_ALIASES[_alias] = _field
del _field, _kind, _lang, _ru_word, _alias


def _normalize_header(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    text = re.sub(r"[^\w\s]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def _normalize_language(value: str | None) -> Optional[str]:
    if value is None:
        return None
    normalized = _normalize_header(value)
    if not normalized:
        return None
    for lang, aliases in SHEET_LANGUAGE_ALIASES.items():
        if normalized in aliases:
            return lang
    return None


def _find_header_row(rows: Sequence[Sequence[object]]) -> Optional[Tuple[int, int, int]]:
    alias_map: dict[str, str] = {}
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_map[alias] = key

    for row_index, row in enumerate(rows):
        header_positions: dict[str, int] = {}
        for col_index, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            target = alias_map.get(normalized)
            if target and target not in header_positions:
                header_positions[target] = col_index
        if "question" in header_positions and "answer" in header_positions:
            return row_index, header_positions["question"], header_positions["answer"]
    return None


def _iter_rows(path: Path, sheet: Optional[str]) -> Iterable[Sequence[object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found. Available: {', '.join(workbook.sheetnames)}")
        worksheet = workbook[sheet]
    else:
        worksheet = workbook.active
    for row in worksheet.iter_rows(values_only=True):
        yield row


def _list_sheet_names(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    return list(workbook.sheetnames)


def _extract_items_from_rows(rows: Sequence[Sequence[object]], limit: Optional[int]) -> List[Tuple[str, str]]:
    if not rows:
        return []

    header = _find_header_row(rows[: min(len(rows), 10)])
    if header:
        header_row, question_idx, answer_idx = header
        data_rows = rows[header_row + 1 :]
    else:
        max_cols = max(len(row) for row in rows)
        counts = [0] * max_cols
        for row in rows:
            for idx in range(min(len(row), max_cols)):
                val = row[idx]
                if val is None:
                    continue
                if str(val).strip():
                    counts[idx] += 1
        if max_cols < 2 or sorted(counts, reverse=True)[:2].count(0) > 0:
            return []
        top_two = sorted(range(max_cols), key=lambda i: (-counts[i], i))[:2]
        question_idx, answer_idx = sorted(top_two)
        data_rows = rows

    items: List[Tuple[str, str]] = []
    seen: set[Tuple[str, str]] = set()
    for row in data_rows:
        if question_idx >= len(row) or answer_idx >= len(row):
            continue
        question_raw = row[question_idx]
        answer_raw = row[answer_idx]
        if question_raw is None or answer_raw is None:
            continue
        question = str(question_raw).strip()
        answer = str(answer_raw).strip()
        if not question or not answer:
            continue
        key = (question, answer)
        if key in seen:
            continue
        seen.add(key)
        items.append(key)
        if limit and len(items) >= limit:
            break
    return items


def _extract_items(path: Path, sheet: Optional[str], limit: Optional[int]) -> List[Tuple[str, str]]:
    return _extract_items_from_rows(list(_iter_rows(path, sheet)), limit)


def _find_wide_header_row(rows: Sequence[Sequence[object]]) -> Optional[Tuple[int, dict[str, int]]]:
    """Detect a wide multilingual header (question_ru / answer_ru / ...).

    Returns (header_row_index, {field: column_index}) or None. Requires at
    least one language to have both its question and answer columns present,
    so plain "question"/"answer" sheets fall through to the legacy parser.
    """
    for row_index, row in enumerate(rows[: min(len(rows), 10)]):
        positions: dict[str, int] = {}
        for col_index, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            field = _WIDE_HEADER_ALIASES.get(normalized)
            if field and field not in positions:
                positions[field] = col_index
        for lang in ("ru", "en", "uz"):
            if f"question_{lang}" in positions and f"answer_{lang}" in positions:
                return row_index, positions
    return None


def _extract_wide_items_from_rows(
    rows: Sequence[Sequence[object]],
    limit: Optional[int],
) -> Optional[List[dict[str, Optional[str]]]]:
    """Parse wide-format rows into multilingual items.

    Returns None when the rows are not in wide format. Rows without a complete
    RU pair are skipped (question_ru/answer_ru are NOT NULL in the DB);
    half-filled pairs in other languages are dropped to None.
    """
    header = _find_wide_header_row(rows)
    if header is None:
        return None
    header_row, positions = header

    items: List[dict[str, Optional[str]]] = []
    for row in rows[header_row + 1 :]:
        item: dict[str, Optional[str]] = {field: None for field in WIDE_FIELDS}
        for field, idx in positions.items():
            if idx < len(row) and row[idx] is not None:
                text = str(row[idx]).strip()
                item[field] = text or None
        for lang in ("ru", "en", "uz"):
            q_key, a_key = f"question_{lang}", f"answer_{lang}"
            if not (item[q_key] and item[a_key]):
                item[q_key] = None
                item[a_key] = None
        if not item["question_ru"]:
            continue
        items.append(item)
        if limit and len(items) >= limit:
            break
    return items


def _extract_multilingual_items(
    path: Path,
    sheet: Optional[str],
    lang: Optional[str],
    limit: Optional[int],
) -> List[dict[str, Optional[str]]]:
    # Wide multilingual layout takes priority: detected by header on the
    # requested (or active) sheet. Legacy layouts (per-language sheets or a
    # single question/answer sheet) fall through below.
    wide = _extract_wide_items_from_rows(list(_iter_rows(path, sheet)), limit)
    if wide is not None:
        return wide

    if sheet:
        rows = _extract_items(path, sheet, limit)
        resolved_lang = _normalize_language(lang) or _normalize_language(sheet) or "ru"
        items: List[dict[str, Optional[str]]] = []
        for q, a in rows:
            item = {
                "question_ru": None,
                "answer_ru": None,
                "question_en": None,
                "answer_en": None,
                "question_uz": None,
                "answer_uz": None,
            }
            item[f"question_{resolved_lang}"] = q
            item[f"answer_{resolved_lang}"] = a
            items.append(item)
        return items

    requested_lang = _normalize_language(lang)
    if requested_lang:
        rows = _extract_items(path, None, limit)
        items: List[dict[str, Optional[str]]] = []
        for q, a in rows:
            item = {
                "question_ru": None,
                "answer_ru": None,
                "question_en": None,
                "answer_en": None,
                "question_uz": None,
                "answer_uz": None,
            }
            item[f"question_{requested_lang}"] = q
            item[f"answer_{requested_lang}"] = a
            items.append(item)
        return items

    all_sheet_names = _list_sheet_names(path)
    detected = [(name, _normalize_language(name)) for name in all_sheet_names]
    detected = [(name, code) for name, code in detected if code]
    if detected:
        per_lang: dict[str, List[Tuple[str, str]]] = {}
        for sheet_name, lang_code in detected:
            per_lang[lang_code] = _extract_items(path, sheet_name, limit)
        row_count = max((len(rows) for rows in per_lang.values()), default=0)
        items: List[dict[str, Optional[str]]] = []
        for idx in range(row_count):
            item = {
                "question_ru": None,
                "answer_ru": None,
                "question_en": None,
                "answer_en": None,
                "question_uz": None,
                "answer_uz": None,
            }
            for lang_code, rows in per_lang.items():
                if idx >= len(rows):
                    continue
                q, a = rows[idx]
                item[f"question_{lang_code}"] = q
                item[f"answer_{lang_code}"] = a
            if any(item.values()):
                items.append(item)
        return items

    rows = _extract_items(path, None, limit)
    return [
        {
            "question_ru": q,
            "answer_ru": a,
            "question_en": None,
            "answer_en": None,
            "question_uz": None,
            "answer_uz": None,
        }
        for q, a in rows
    ]


def _extract_items_from_csv(path: Path, limit: Optional[int]) -> List[dict[str, Optional[str]]]:
    """Parse a FAQ csv: wide multilingual layout, or a legacy two-column one."""
    import csv

    # utf-8-sig strips the BOM our own export writes for Excel.
    with path.open(newline="", encoding="utf-8-sig") as fh:
        rows = [tuple(row) for row in csv.reader(fh)]

    wide = _extract_wide_items_from_rows(rows, limit)
    if wide is not None:
        return wide
    return [
        {
            "question_ru": q,
            "answer_ru": a,
            "question_en": None,
            "answer_en": None,
            "question_uz": None,
            "answer_uz": None,
        }
        for q, a in _extract_items_from_rows(rows, limit)
    ]


def _extract_items_from_json(path: Path, limit: Optional[int]) -> List[dict[str, Optional[str]]]:
    """Parse a FAQ json export: either the /admin/seed payload or a bare list."""
    import json

    payload = json.loads(path.read_text(encoding="utf-8"))
    raw_items = payload.get("items") if isinstance(payload, dict) else payload
    if not isinstance(raw_items, list):
        raise ValueError("JSON должен быть списком записей или объектом с ключом 'items'.")

    items: List[dict[str, Optional[str]]] = []
    for raw in raw_items:
        if not isinstance(raw, dict):
            continue
        item: dict[str, Optional[str]] = {}
        for field in WIDE_FIELDS:
            value = raw.get(field)
            text = str(value).strip() if value is not None else ""
            item[field] = text or None
        for lang in ("en", "uz"):
            q_key, a_key = f"question_{lang}", f"answer_{lang}"
            if not (item[q_key] and item[a_key]):
                item[q_key] = None
                item[a_key] = None
        if not (item["question_ru"] and item["answer_ru"]):
            continue
        items.append(item)
        if limit and len(items) >= limit:
            break
    return items


def _extract_multilingual_items_any(path: Path, limit: Optional[int] = None) -> List[dict[str, Optional[str]]]:
    """Dispatch FAQ extraction by file extension: xlsx / csv / json."""
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _extract_items_from_json(path, limit)
    if suffix == ".csv":
        return _extract_items_from_csv(path, limit)
    return _extract_multilingual_items(path, None, None, limit)


async def _attach_embeddings(items: List[dict[str, Optional[str]]]) -> None:
    """Batch-embed all questions across languages and attach vectors to *items*.

    Three batched OpenAI calls (one per language). On failure any individual
    embedding stays None — the row is still inserted, the SQLAlchemy event
    listener will not retry (its before_insert path skips columns that already
    exist or are explicitly None for missing questions). Backfill via the
    admin UI handles those rows later.
    """
    import logging

    from app.utils.embeddings import embed_texts

    logger = logging.getLogger(__name__)

    for lang in ("ru", "en", "uz"):
        q_field = f"question_{lang}"
        emb_field = f"embedding_{lang}"
        # Pair (index, text) so we can write back after the batch.
        pairs = [(i, item.get(q_field)) for i, item in enumerate(items)]
        texts = [t or "" for _, t in pairs]
        if not any(texts):
            continue
        try:
            vectors = await embed_texts(texts)
        except Exception as exc:
            logger.warning("seed embed batch failed for %s: %s", lang, exc)
            continue
        for (idx, q_text), vec in zip(pairs, vectors):
            if q_text and vec is not None:
                items[idx][emb_field] = vec


async def _import_items(items: List[dict[str, Optional[str]]], replace: bool, dry_run: bool) -> None:
    if dry_run:
        return
    await _attach_embeddings(items)
    async with get_session() as session:
        if replace:
            await session.execute(delete(FaqItem))
        session.add_all([FaqItem(**item) for item in items])
        await session.commit()
