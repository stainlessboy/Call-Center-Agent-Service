"""Seed filials / sales offices / sales points from three uploaded xlsx files.

Sales offices and sales points reference their parent filial by a short name
that does not always exactly match the filial's full name — we use fuzzy
matching (difflib) with normalized names to resolve the link.
"""
from __future__ import annotations

import difflib
import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy import delete

from app.db.models import Filial, SalesOffice, SalesPoint
from app.db.session import get_session

log = logging.getLogger(__name__)

SHEET_RU = "Рус"
SHEET_UZ = "Узб (латиница)"

_STRIP_PREFIXES_RE = re.compile(r"\b(ЦБУ|BXM|БХМ)\b", re.IGNORECASE)
_NON_WORD_RE = re.compile(r"[^\w\s]", re.UNICODE)
_WS_RE = re.compile(r"\s+")


def _normalize(s: Optional[str]) -> str:
    if not s:
        return ""
    s = _STRIP_PREFIXES_RE.sub(" ", s)
    s = _NON_WORD_RE.sub(" ", s)
    s = _WS_RE.sub(" ", s).strip().lower()
    return s


def _read_sheet_rows(path: Path, sheet_name: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _strip(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _load_filials(path: Path) -> list[dict]:
    ru_rows = _read_sheet_rows(path, SHEET_RU)
    uz_rows = _read_sheet_rows(path, SHEET_UZ)

    def _data_rows(rows):
        out = []
        for r in rows:
            if not any(c for c in r):
                continue
            if r[0] == "№":
                continue
            out.append(r)
        return out

    ru = _data_rows(ru_rows)
    uz = _data_rows(uz_rows)

    items: list[dict] = []
    for idx, ru_row in enumerate(ru):
        uz_row = uz[idx] if idx < len(uz) else (None,) * len(ru_row)
        items.append({
            "name_ru": _strip(ru_row[1]),
            "name_uz": _strip(uz_row[1]),
            "address_ru": _strip(ru_row[2]),
            "address_uz": _strip(uz_row[2]),
            "landmark_ru": _strip(ru_row[3]),
            "landmark_uz": _strip(uz_row[3]),
            "location_url": _strip(ru_row[4]) or _strip(uz_row[4]),
        })
    log.info("filials loaded: %d rows", len(items))
    return items


def _load_sales_offices(path: Path) -> list[dict]:
    ru_rows = _read_sheet_rows(path, SHEET_RU)
    uz_rows = _read_sheet_rows(path, SHEET_UZ)

    def _data_rows(rows):
        out = []
        for r in rows:
            if not any(c for c in r):
                continue
            if r[0] and str(r[0]).startswith(("ЦБУ", "BXM", "БХМ")) and ("филиал" in str(r[0]).lower() or "filial" in str(r[0]).lower()):
                continue
            out.append(r)
        return out

    ru = _data_rows(ru_rows)
    uz = _data_rows(uz_rows)

    items: list[dict] = []
    for idx, ru_row in enumerate(ru):
        uz_row = uz[idx] if idx < len(uz) else (None,) * len(ru_row)
        items.append({
            "parent_ref_ru": _strip(ru_row[0]),
            "name_ru": _strip(ru_row[1]),
            "name_uz": _strip(uz_row[1]),
            "region_ru": _strip(ru_row[2]),
            "region_uz": _strip(uz_row[2]),
            "address_ru": _strip(ru_row[3]),
            "address_uz": _strip(uz_row[3]),
        })
    log.info("sales_offices loaded: %d rows", len(items))
    return items


def _load_sales_points(path: Path) -> list[dict]:
    ru_rows = _read_sheet_rows(path, SHEET_RU)
    uz_rows = _read_sheet_rows(path, SHEET_UZ)

    def _data_rows(rows):
        out = []
        for r in rows:
            if not any(c for c in r):
                continue
            if r[0] == "№":
                continue
            out.append(r)
        return out

    ru = _data_rows(ru_rows)
    uz = _data_rows(uz_rows)

    items: list[dict] = []
    for idx, ru_row in enumerate(ru):
        uz_row = uz[idx] if idx < len(uz) else (None,) * len(ru_row)
        items.append({
            "parent_ref_ru": _strip(ru_row[1]),
            "name_ru": _strip(ru_row[2]),
            "name_uz": _strip(uz_row[2]),
            "address_ru": _strip(ru_row[3]),
            "address_uz": _strip(uz_row[3]),
        })
    log.info("sales_points loaded: %d rows", len(items))
    return items


def _resolve_parent(
    ref: Optional[str], filial_index: dict[str, int], cutoff: float = 0.6
) -> Optional[int]:
    if not ref:
        return None
    key = _normalize(ref)
    if key in filial_index:
        return filial_index[key]
    matches = difflib.get_close_matches(key, list(filial_index.keys()), n=1, cutoff=cutoff)
    if matches:
        return filial_index[matches[0]]
    log.warning("no parent filial match for ref=%r (normalized=%r)", ref, key)
    return None


async def _seed(
    filials_path: Path,
    offices_path: Path,
    points_path: Path,
    replace: bool,
) -> None:
    filials_data = _load_filials(filials_path)
    offices_data = _load_sales_offices(offices_path)
    points_data = _load_sales_points(points_path)

    async with get_session() as session:
        if replace:
            log.info("--replace: deleting all rows from filials/sales_offices/sales_points")
            await session.execute(delete(SalesPoint))
            await session.execute(delete(SalesOffice))
            await session.execute(delete(Filial))
            await session.flush()

        filial_objs: list[Filial] = []
        for d in filials_data:
            if not d.get("name_ru") or not d.get("address_ru"):
                continue
            filial_objs.append(Filial(
                name_ru=d["name_ru"],
                name_uz=d.get("name_uz"),
                address_ru=d["address_ru"],
                address_uz=d.get("address_uz"),
                landmark_ru=d.get("landmark_ru"),
                landmark_uz=d.get("landmark_uz"),
                location_url=d.get("location_url"),
            ))
        session.add_all(filial_objs)
        await session.flush()

        filial_index: dict[str, int] = {}
        for f in filial_objs:
            filial_index[_normalize(f.name_ru)] = f.id
            if f.name_uz:
                filial_index[_normalize(f.name_uz)] = f.id

        log.info("inserted %d filials, building parent-index with %d keys",
                 len(filial_objs), len(filial_index))

        office_objs: list[SalesOffice] = []
        for d in offices_data:
            if not d.get("name_ru") or not d.get("address_ru"):
                continue
            parent_id = _resolve_parent(d.get("parent_ref_ru"), filial_index)
            office_objs.append(SalesOffice(
                name_ru=d["name_ru"],
                name_uz=d.get("name_uz"),
                region_ru=d.get("region_ru"),
                region_uz=d.get("region_uz"),
                address_ru=d["address_ru"],
                address_uz=d.get("address_uz"),
                parent_filial_id=parent_id,
            ))
        session.add_all(office_objs)
        log.info("inserted %d sales_offices", len(office_objs))

        point_objs: list[SalesPoint] = []
        for d in points_data:
            if not d.get("name_ru") or not d.get("address_ru"):
                continue
            parent_id = _resolve_parent(d.get("parent_ref_ru"), filial_index)
            point_objs.append(SalesPoint(
                name_ru=d["name_ru"],
                name_uz=d.get("name_uz"),
                address_ru=d["address_ru"],
                address_uz=d.get("address_uz"),
                parent_filial_id=parent_id,
            ))
        session.add_all(point_objs)
        log.info("inserted %d sales_points", len(point_objs))

        await session.commit()
