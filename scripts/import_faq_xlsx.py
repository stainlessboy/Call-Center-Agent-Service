#!/usr/bin/env python3
from __future__ import annotations

import argparse
import asyncio
import re
import sys
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook
from sqlalchemy import delete

from app.db.models import FaqItem
from app.db.session import get_session

HEADER_ALIASES = {
    "question": {
        "question",
        "questions",
        "q",
        "faq question",
        "вопрос",
        "вопросы",
        "вопрос/проблема",
        "вопросы/проблемы",
        "вопросы по кредитам",
        "вопросы по кредитам?",
    },
    "answer": {
        "answer",
        "answers",
        "a",
        "response",
        "ответ",
        "ответы",
        "ответ/решение",
        "ответы/решения",
    },
}

SHEET_LANGUAGE_ALIASES = {
    "ru": {"ru", "russian", "русский", "рус", "российский"},
    "en": {"en", "eng", "english", "английский", "англ"},
    "uz": {"uz", "uzb", "uzbek", "uzbekskiy", "узбекский", "узб"},
}


def _normalize_header(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    text = re.sub(r"[^\w\s]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def _normalize_language(value: str | None) -> Optional[str]:
    if value is None:
        return None
    normalized = _normalize_header(value)
    if not normalized:
        return None
    for lang, aliases in SHEET_LANGUAGE_ALIASES.items():
        if normalized in aliases:
            return lang
    return None


def _find_header_row(rows: Sequence[Sequence[object]]) -> Optional[Tuple[int, int, int]]:
    alias_map: dict[str, str] = {}
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_map[alias] = key

    for row_index, row in enumerate(rows):
        header_positions: dict[str, int] = {}
        for col_index, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            target = alias_map.get(normalized)
            if target and target not in header_positions:
                header_positions[target] = col_index
        if "question" in header_positions and "answer" in header_positions:
            return row_index, header_positions["question"], header_positions["answer"]
    return None


def _iter_rows(path: Path, sheet: Optional[str]) -> Iterable[Sequence[object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found. Available: {', '.join(workbook.sheetnames)}")
        worksheet = workbook[sheet]
    else:
        worksheet = workbook.active
    for row in worksheet.iter_rows(values_only=True):
        yield row


def _list_sheet_names(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    return list(workbook.sheetnames)


def _extract_items_from_rows(rows: Sequence[Sequence[object]], limit: Optional[int]) -> List[Tuple[str, str]]:
    if not rows:
        return []

    header = _find_header_row(rows[: min(len(rows), 10)])
    if header:
        header_row, question_idx, answer_idx = header
        data_rows = rows[header_row + 1 :]
    else:
        max_cols = max(len(row) for row in rows)
        counts = [0] * max_cols
        for row in rows:
            for idx in range(min(len(row), max_cols)):
                val = row[idx]
                if val is None:
                    continue
                if str(val).strip():
                    counts[idx] += 1
        if max_cols < 2 or sorted(counts, reverse=True)[:2].count(0) > 0:
            return []
        top_two = sorted(range(max_cols), key=lambda i: (-counts[i], i))[:2]
        question_idx, answer_idx = sorted(top_two)
        data_rows = rows

    items: List[Tuple[str, str]] = []
    seen: set[Tuple[str, str]] = set()
    for row in data_rows:
        if question_idx >= len(row) or answer_idx >= len(row):
            continue
        question_raw = row[question_idx]
        answer_raw = row[answer_idx]
        if question_raw is None or answer_raw is None:
            continue
        question = str(question_raw).strip()
        answer = str(answer_raw).strip()
        if not question or not answer:
            continue
        key = (question, answer)
        if key in seen:
            continue
        seen.add(key)
        items.append(key)
        if limit and len(items) >= limit:
            break
    return items


def _extract_items(path: Path, sheet: Optional[str], limit: Optional[int]) -> List[Tuple[str, str]]:
    return _extract_items_from_rows(list(_iter_rows(path, sheet)), limit)


def _extract_multilingual_items(
    path: Path,
    sheet: Optional[str],
    lang: Optional[str],
    limit: Optional[int],
) -> List[dict[str, Optional[str]]]:
    if sheet:
        rows = _extract_items(path, sheet, limit)
        resolved_lang = _normalize_language(lang) or _normalize_language(sheet) or "ru"
        items: List[dict[str, Optional[str]]] = []
        for q, a in rows:
            item = {
                "question_ru": None,
                "answer_ru": None,
                "question_en": None,
                "answer_en": None,
                "question_uz": None,
                "answer_uz": None,
            }
            item[f"question_{resolved_lang}"] = q
            item[f"answer_{resolved_lang}"] = a
            items.append(item)
        return items

    requested_lang = _normalize_language(lang)
    if requested_lang:
        rows = _extract_items(path, None, limit)
        items: List[dict[str, Optional[str]]] = []
        for q, a in rows:
            item = {
                "question_ru": None,
                "answer_ru": None,
                "question_en": None,
                "answer_en": None,
                "question_uz": None,
                "answer_uz": None,
            }
            item[f"question_{requested_lang}"] = q
            item[f"answer_{requested_lang}"] = a
            items.append(item)
        return items

    all_sheet_names = _list_sheet_names(path)
    detected = [(name, _normalize_language(name)) for name in all_sheet_names]
    detected = [(name, code) for name, code in detected if code]
    if detected:
        per_lang: dict[str, List[Tuple[str, str]]] = {}
        for sheet_name, lang_code in detected:
            per_lang[lang_code] = _extract_items(path, sheet_name, limit)
        row_count = max((len(rows) for rows in per_lang.values()), default=0)
        items: List[dict[str, Optional[str]]] = []
        for idx in range(row_count):
            item = {
                "question_ru": None,
                "answer_ru": None,
                "question_en": None,
                "answer_en": None,
                "question_uz": None,
                "answer_uz": None,
            }
            for lang_code, rows in per_lang.items():
                if idx >= len(rows):
                    continue
                q, a = rows[idx]
                item[f"question_{lang_code}"] = q
                item[f"answer_{lang_code}"] = a
            if any(item.values()):
                items.append(item)
        return items

    rows = _extract_items(path, None, limit)
    return [
        {
            "question_ru": q,
            "answer_ru": a,
            "question_en": None,
            "answer_en": None,
            "question_uz": None,
            "answer_uz": None,
        }
        for q, a in rows
    ]


async def _import_items(items: List[dict[str, Optional[str]]], replace: bool, dry_run: bool) -> None:
    if dry_run:
        return
    async with get_session() as session:
        if replace:
            await session.execute(delete(FaqItem))
        session.add_all([FaqItem(**item) for item in items])
        await session.commit()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import FAQ items from an .xlsx file into the database.")
    parser.add_argument("path", type=Path, help="Path to FAQ .xlsx file")
    parser.add_argument("--sheet", type=str, help="Sheet name (defaults to auto-detect all language sheets)")
    parser.add_argument("--lang", type=str, choices=["ru", "en", "uz"], help="Language code for single-sheet import")
    parser.add_argument("--replace", action="store_true", help="Delete existing FAQ items before import")
    parser.add_argument("--limit", type=int, help="Import only first N rows per sheet (for testing)")
    parser.add_argument("--dry-run", action="store_true", help="Parse and report rows without writing to DB")
    args = parser.parse_args()

    if not args.path.exists():
        raise SystemExit(f"File not found: {args.path}")

    items = _extract_multilingual_items(args.path, args.sheet, args.lang, args.limit)
    print(f"Parsed {len(items)} FAQ rows from {args.path}")
    if items:
        filled_counts = {
            "ru": sum(1 for item in items if item.get("question_ru") and item.get("answer_ru")),
            "en": sum(1 for item in items if item.get("question_en") and item.get("answer_en")),
            "uz": sum(1 for item in items if item.get("question_uz") and item.get("answer_uz")),
        }
        print("Filled languages:", ", ".join(f"{lang}={count}" for lang, count in sorted(filled_counts.items())))
        print("Sample:")
        for item in items[:3]:
            print(f"- [ru] Q: {item.get('question_ru')}")
            print(f"      A: {item.get('answer_ru')}")
            if item.get("question_en"):
                print(f"  [en] Q: {item.get('question_en')}")
                print(f"      A: {item.get('answer_en')}")
            if item.get("question_uz"):
                print(f"  [uz] Q: {item.get('question_uz')}")
                print(f"      A: {item.get('answer_uz')}")
    if args.dry_run:
        print("Dry-run mode: no changes were written to the database.")
        return

    asyncio.run(_import_items(items, args.replace, args.dry_run))
    print("Import complete.")


if __name__ == "__main__":
    main()
