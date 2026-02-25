#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import load_workbook

SECTION_FILENAMES: Dict[tuple[str, str], str] = {
    ("credit_products", "Микрозайм"): "credit_microloan.json",
    ("credit_products", "Ипотека"): "credit_mortgage.json",
    ("credit_products", "Автокредит"): "credit_auto_loan.json",
    ("credit_products", "Образовательный"): "credit_education.json",
    ("noncredit_products", "Вклады"): "noncredit_deposits.json",
    ("noncredit_products", "Карты"): "noncredit_cards.json",
}


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _section_title(row: Sequence[object]) -> Optional[str]:
    if not row:
        return None
    first = row[0]
    if _is_blank(first):
        return None
    if any(not _is_blank(cell) for cell in row[1:]):
        return None
    return str(first).strip() or None


def _clean_row(row: Sequence[object], max_cols: int) -> List[Any]:
    cleaned: List[Any] = []
    for i in range(max_cols):
        value = row[i] if i < len(row) else None
        if isinstance(value, str):
            value = value.strip()
        cleaned.append(value)
    return cleaned


def _normalize_rows(rows: List[List[Any]]) -> List[List[Any]]:
    if not rows:
        return []
    max_cols = max(len(r) for r in rows)
    last: List[Any] = [None] * max_cols
    normalized: List[List[Any]] = []
    for row in rows:
        filled: List[Any] = []
        for idx in range(max_cols):
            value = row[idx] if idx < len(row) else None
            if _is_blank(value):
                value = last[idx]
            else:
                last[idx] = value
            filled.append(value)
        normalized.append(filled)
    return normalized


def _parse_sheet(path: Path, sheet_name: str) -> Dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")

    ws = wb[sheet_name]
    rows_raw = list(ws.iter_rows(values_only=True))
    sections: Dict[str, Dict[str, Any]] = {}

    i = 0
    while i < len(rows_raw):
        row = rows_raw[i]
        title = _section_title(row)
        if not title:
            i += 1
            continue

        # find header row
        i += 1
        while i < len(rows_raw) and all(_is_blank(v) for v in rows_raw[i]):
            i += 1
        if i >= len(rows_raw):
            break

        header_row = rows_raw[i]
        max_cols = max(len(header_row), 1)
        header = _clean_row(header_row, max_cols)
        i += 1

        data_rows: List[List[Any]] = []
        while i < len(rows_raw):
            next_row = rows_raw[i]
            if _section_title(next_row):
                break
            if all(_is_blank(v) for v in next_row):
                i += 1
                continue
            data_rows.append(_clean_row(next_row, max_cols))
            i += 1

        sections[title] = {
            "header": header,
            "rows_raw": data_rows,
            "rows_normalized": _normalize_rows(data_rows),
        }

    return {
        "sheet": sheet_name,
        "sections": sections,
    }


def _fallback_section_filename(sheet_key: str, section_name: str, order: int) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", section_name.strip().lower())
    normalized = normalized.strip("_")
    if not normalized:
        normalized = f"section_{order:02d}"
    prefix = "credit" if sheet_key == "credit_products" else "noncredit"
    return f"{prefix}_{normalized}.json"


def _build_split_manifest(
    workbook: Dict[str, Dict[str, Any]],
    split_dir: Path,
    manifest_base_dir: Path,
) -> Dict[str, Dict[str, str]]:
    split_dir.mkdir(parents=True, exist_ok=True)
    for stale in split_dir.glob("*.json"):
        stale.unlink()

    manifest: Dict[str, Dict[str, str]] = {
        "credit_products": {},
        "noncredit_products": {},
    }

    for sheet_key, sheet_payload in workbook.items():
        sections = sheet_payload.get("sections")
        if not isinstance(sections, dict):
            continue
        for order, (section_name, section_payload) in enumerate(sections.items(), start=1):
            if not isinstance(section_name, str) or not isinstance(section_payload, dict):
                continue
            filename = SECTION_FILENAMES.get(
                (sheet_key, section_name),
                _fallback_section_filename(sheet_key, section_name, order),
            )
            out_path = split_dir / filename
            payload = {
                "sheet_key": sheet_key,
                "sheet_name": sheet_payload.get("sheet"),
                "section_name": section_name,
                "header": section_payload.get("header") or [],
                "rows_raw": section_payload.get("rows_raw") or [],
                "rows_normalized": section_payload.get("rows_normalized") or [],
            }
            out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            try:
                section_path = out_path.relative_to(manifest_base_dir).as_posix()
            except ValueError:
                section_path = str(out_path)
            manifest.setdefault(sheet_key, {})[section_name] = section_path
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Export AI CHAT INFO.xlsx into JSON.")
    parser.add_argument("path", type=Path, help="Path to AI CHAT INFO.xlsx")
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("app/data/ai_chat_info.json"),
        help="Output JSON path (default: app/data/ai_chat_info.json)",
    )
    parser.add_argument(
        "--split-dir",
        type=Path,
        default=Path("app/data/ai_chat_info"),
        help="Directory for split section JSON files (default: app/data/ai_chat_info)",
    )
    args = parser.parse_args()

    if not args.path.exists():
        raise SystemExit(f"File not found: {args.path}")

    workbook = {
        "credit_products": _parse_sheet(args.path, "Кредитные продукты"),
        "noncredit_products": _parse_sheet(args.path, "Некредитные продукты"),
    }
    split_manifest = _build_split_manifest(workbook, args.split_dir, args.out.parent)
    result = {
        "source_file": str(args.path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "format_version": 2,
        "layout": split_manifest,
    }

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {args.out}")
    print(f"Wrote split sections to {args.split_dir}")


if __name__ == "__main__":
    main()
